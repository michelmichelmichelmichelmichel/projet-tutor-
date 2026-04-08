#!/usr/bin/env python3
"""
Parse Romanian INSSE Tourism Excel files (TUR102C, TUR103F, TUR104H, TUR105H)
and generate a structured JSON file grouped by County.

- TUR102C: Annual capacity (nombre d'hébergements par an)
- TUR103F: Monthly bed capacity (places-jours par mois)
- TUR104H: Monthly tourist arrivals (arrivées par mois)
- TUR105H: Monthly overnight stays (nuitées par mois)

Output: romania_tourism_data.json
"""

import openpyxl
import json
import os
import re
from collections import defaultdict

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

FILES = {
    "TUR102C": "TUR102C - Existing touristic accommodation capacity by type of establishment, counties and localities (1) (nombre hotels par an).xlsx",
    "TUR103F": "TUR103F - Touristic accommodation monthly capacity in function by type of establishment, counties and localities (1).xlsx",
    "TUR104H": "TUR104H - Arrivals of tourists accommodated in the structure of tourists reception by type of establishment, by counties and localities, monthly (1).xlsx",
    "TUR105H": "TUR105H - Staying overnight in the establishments of touristic reception by counties and localities, monthly (1).xlsx",
}

# Establishment types to skip (metadata rows, not real data)
SKIP_TYPES = {
    "© 1998 - 2018 INSTITUTUL NATIONAL DE STATISTICA",
}

# Valid establishment types
VALID_TYPES = {
    "Total", "Hotels", "Motels", "Inns", "Hostels",
    "Touristic villas", "Touristic chalets", "Touristic boarding houses",
    "Agroturistic boarding houses", "Campings", "Bungalows",
    "Holiday villages", "Apartament hotels", "Houselet type unit",
    "School and pre-school camps", "Ships accommodation spaces",
    "Apartments and rooms for rent", "Touristic halting places",
}


def parse_month_header(header):
    """Parse a month header like 'January 2023' -> (2023, 1)"""
    months = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8,
        "September": 9, "October": 10, "November": 11, "December": 12,
    }
    parts = header.strip().split()
    if len(parts) == 2 and parts[0] in months:
        return (int(parts[1]), months[parts[0]])
    return None


def parse_year_header(header):
    """Parse a year header like 'Year 2023' -> 2023"""
    match = re.match(r"Year\s+(\d{4})", header.strip())
    if match:
        return int(match.group(1))
    return None


def clean_value(val):
    """Clean a cell value: handle missing data markers, convert to number."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in (":", "c", "-", "", "..."):
            return None
        try:
            return float(val.replace(",", "."))
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return val
    return None


def read_monthly_file(filepath):
    """
    Read a monthly INSSE Excel file (TUR103F, TUR104H, TUR105H).
    Returns: dict[county][establishment_type] = {(year, month): value}
    And the list of (year, month) tuples found.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse date headers from row 3
    row3 = [c.value for c in list(ws.iter_rows(min_row=3, max_row=3))[0]]
    date_columns = {}  # col_index -> (year, month)
    for i, val in enumerate(row3):
        if val and isinstance(val, str):
            parsed = parse_month_header(val)
            if parsed:
                date_columns[i] = parsed

    # Parse data rows starting from row 6
    data = defaultdict(lambda: defaultdict(dict))
    for row in ws.iter_rows(min_row=6, values_only=True):
        est_type = row[0]
        county = row[1]
        # locality = row[2]  # Always "TOTAL" in our data

        if not est_type or not county:
            continue
        if est_type in SKIP_TYPES or est_type not in VALID_TYPES:
            continue

        for col_idx, date_key in date_columns.items():
            if col_idx < len(row):
                val = clean_value(row[col_idx])
                if val is not None:
                    data[county][est_type][date_key] = val

    wb.close()
    periods = sorted(date_columns.values())
    return data, periods


def read_annual_file(filepath):
    """
    Read the annual INSSE Excel file (TUR102C).
    Returns: dict[county][establishment_type] = {year: value}
    And the list of years found.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Parse year headers from row 3
    row3 = [c.value for c in list(ws.iter_rows(min_row=3, max_row=3))[0]]
    year_columns = {}  # col_index -> year
    for i, val in enumerate(row3):
        if val and isinstance(val, str):
            parsed = parse_year_header(val)
            if parsed:
                year_columns[i] = parsed

    # Parse data rows starting from row 6
    data = defaultdict(lambda: defaultdict(dict))
    for row in ws.iter_rows(min_row=6, values_only=True):
        est_type = row[0]
        county = row[1]

        if not est_type or not county:
            continue
        if est_type in SKIP_TYPES or est_type not in VALID_TYPES:
            continue

        for col_idx, year in year_columns.items():
            if col_idx < len(row):
                val = clean_value(row[col_idx])
                if val is not None:
                    data[county][est_type][year] = val

    wb.close()
    years = sorted(year_columns.values())
    return data, years


def find_common_period(periods_103, periods_104, periods_105):
    """Find the intersection of monthly periods across the 3 monthly datasets."""
    set_103 = set(periods_103)
    set_104 = set(periods_104)
    set_105 = set(periods_105)
    common = set_103 & set_104 & set_105
    return sorted(common)


def build_structured_json(data_102, years_102, data_103, data_104, data_105, common_months):
    """
    Build the final structured JSON.
    
    Structure:
    {
        "metadata": { ... },
        "counties": {
            "CountyName": {
                "annual_capacity": {
                    "total": { "2024": value, ... },
                    "by_type": {
                        "Hotels": { "2024": value, ... },
                        ...
                    }
                },
                "monthly_data": {
                    "2024-01": {
                        "bed_capacity": { "total": value, "by_type": {...} },
                        "arrivals": { "total": value, "by_type": {...} },
                        "overnight_stays": { "total": value, "by_type": {...} }
                    },
                    ...
                }
            }
        }
    }
    """
    # Get all unique counties
    all_counties = set()
    for d in [data_102, data_103, data_104, data_105]:
        all_counties.update(d.keys())
    all_counties = sorted(all_counties)

    # Determine the years covered by common months
    common_years = sorted(set(y for y, m in common_months))

    result = {
        "metadata": {
            "source": "INSSE Romania - TEMPO Online",
            "datasets": {
                "TUR102C": {
                    "name": "Existing touristic accommodation capacity",
                    "unit": "Number of places",
                    "granularity": "annual",
                    "years": years_102,
                },
                "TUR103F": {
                    "name": "Touristic accommodation monthly capacity in function",
                    "unit": "Places-days",
                    "granularity": "monthly",
                },
                "TUR104H": {
                    "name": "Arrivals of tourists accommodated",
                    "unit": "Persons",
                    "granularity": "monthly",
                },
                "TUR105H": {
                    "name": "Overnight stays in touristic establishments",
                    "unit": "Number",
                    "granularity": "monthly",
                },
            },
            "aligned_period": {
                "monthly_start": f"{common_months[0][0]}-{common_months[0][1]:02d}",
                "monthly_end": f"{common_months[-1][0]}-{common_months[-1][1]:02d}",
                "months_count": len(common_months),
                "annual_years": common_years,
            },
            "counties_count": len(all_counties),
        },
        "counties": {},
    }

    for county in all_counties:
        county_data = {
            "annual_capacity": {
                "total": {},
                "by_type": {},
            },
            "monthly_data": {},
        }

        # Annual capacity (TUR102C) - only for years covered by common months
        if county in data_102:
            for est_type, year_vals in data_102[county].items():
                for year in years_102:
                    if year in year_vals:
                        year_str = str(year)
                        if est_type == "Total":
                            county_data["annual_capacity"]["total"][year_str] = year_vals[year]
                        else:
                            if est_type not in county_data["annual_capacity"]["by_type"]:
                                county_data["annual_capacity"]["by_type"][est_type] = {}
                            county_data["annual_capacity"]["by_type"][est_type][year_str] = year_vals[year]

        # Monthly data - aligned on common period
        for year, month in common_months:
            month_key = f"{year}-{month:02d}"
            month_entry = {
                "bed_capacity": {"total": None, "by_type": {}},
                "arrivals": {"total": None, "by_type": {}},
                "overnight_stays": {"total": None, "by_type": {}},
            }

            # TUR103F - Bed capacity
            if county in data_103:
                for est_type, vals in data_103[county].items():
                    if (year, month) in vals:
                        v = vals[(year, month)]
                        if est_type == "Total":
                            month_entry["bed_capacity"]["total"] = v
                        else:
                            month_entry["bed_capacity"]["by_type"][est_type] = v

            # TUR104H - Arrivals
            if county in data_104:
                for est_type, vals in data_104[county].items():
                    if (year, month) in vals:
                        v = vals[(year, month)]
                        if est_type == "Total":
                            month_entry["arrivals"]["total"] = v
                        else:
                            month_entry["arrivals"]["by_type"][est_type] = v

            # TUR105H - Overnight stays
            if county in data_105:
                for est_type, vals in data_105[county].items():
                    if (year, month) in vals:
                        v = vals[(year, month)]
                        if est_type == "Total":
                            month_entry["overnight_stays"]["total"] = v
                        else:
                            month_entry["overnight_stays"]["by_type"][est_type] = v

            county_data["monthly_data"][month_key] = month_entry

        result["counties"][county] = county_data

    return result


def main():
    print("=" * 60)
    print("INSSE Romania Tourism Data Parser")
    print("=" * 60)

    # 1. Read TUR102C (Annual capacity)
    print("\n[1/4] Reading TUR102C (Annual capacity)...")
    path_102 = os.path.join(DATA_DIR, FILES["TUR102C"])
    data_102, years_102 = read_annual_file(path_102)
    print(f"  → {len(data_102)} counties, years: {years_102}")

    # 2. Read TUR103F (Monthly bed capacity)
    print("\n[2/4] Reading TUR103F (Monthly bed capacity)...")
    path_103 = os.path.join(DATA_DIR, FILES["TUR103F"])
    data_103, periods_103 = read_monthly_file(path_103)
    print(f"  → {len(data_103)} counties, {len(periods_103)} months: {periods_103[0]} to {periods_103[-1]}")

    # 3. Read TUR104H (Monthly arrivals)
    print("\n[3/4] Reading TUR104H (Monthly arrivals)...")
    path_104 = os.path.join(DATA_DIR, FILES["TUR104H"])
    data_104, periods_104 = read_monthly_file(path_104)
    print(f"  → {len(data_104)} counties, {len(periods_104)} months: {periods_104[0]} to {periods_104[-1]}")

    # 4. Read TUR105H (Monthly overnight stays)
    print("\n[4/4] Reading TUR105H (Monthly overnight stays)...")
    path_105 = os.path.join(DATA_DIR, FILES["TUR105H"])
    data_105, periods_105 = read_monthly_file(path_105)
    print(f"  → {len(data_105)} counties, {len(periods_105)} months: {periods_105[0]} to {periods_105[-1]}")

    # 5. Find common period
    print("\n[Align] Finding common monthly period...")
    common_months = find_common_period(periods_103, periods_104, periods_105)
    print(f"  → Common period: {common_months[0]} to {common_months[-1]} ({len(common_months)} months)")

    # 6. Build JSON
    print("\n[Build] Generating structured JSON...")
    result = build_structured_json(data_102, years_102, data_103, data_104, data_105, common_months)

    # 7. Write output
    output_path = os.path.join(DATA_DIR, "romania_tourism_data.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    file_size = os.path.getsize(output_path)
    print(f"\n{'='*60}")
    print(f"✅ Output: {output_path}")
    print(f"   Size: {file_size / 1024:.1f} KB")
    print(f"   Counties: {result['metadata']['counties_count']}")
    print(f"   Aligned period: {result['metadata']['aligned_period']['monthly_start']} → {result['metadata']['aligned_period']['monthly_end']}")
    print(f"   Monthly data points: {result['metadata']['aligned_period']['months_count']} months × {result['metadata']['counties_count']} counties")

    # Summary stats
    print(f"\n--- Sample data for Brasov ---")
    brasov = result["counties"].get("Brasov", {})
    if brasov:
        cap = brasov["annual_capacity"]["total"]
        print(f"  Annual capacity (Total): {cap}")
        first_month = list(brasov["monthly_data"].keys())[0]
        md = brasov["monthly_data"][first_month]
        print(f"  {first_month}:")
        print(f"    Bed capacity (total): {md['bed_capacity']['total']}")
        print(f"    Arrivals (total): {md['arrivals']['total']}")
        print(f"    Overnight stays (total): {md['overnight_stays']['total']}")


if __name__ == "__main__":
    main()
